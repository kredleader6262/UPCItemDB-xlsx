# Importing necessary libraries
import requests
from openpyxl import Workbook, load_workbook
import os
import time
import sys
import configparser
from tqdm import tqdm

# Load user key from config file
def load_user_key():
    filename = "config.ini"
    config = configparser.ConfigParser()
    config.read(filename)
    user_key = config["UPCITEMDB"]["user_key"]
    if user_key:
        print("User key loaded")
        return user_key
    else:
        print("User key not found. Using free tier")
        return None
def load_skip_duplicates():
    filename = "config.ini"
    config = configparser.ConfigParser()
    config.read(filename)
    skip_duplicates = config["UPCITEMDB"].getboolean("skip_duplicates")
    if skip_duplicates:
        return True
    else:
        return False

# Function to check and load or create workbook
def load_or_create_workbook(filename):
    if os.path.exists(filename):
        print(f"Loading existing workbook: {filename}")
        return load_workbook(filename)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Master"
        # Extending header for the master sheet with price metrics
        headers = ["EAN", "Title", "UPC", "GTIN", "ASIN", "Description", "Brand", "Model", "Dimension", "Weight", "Category", "Currency", "Lowest Recorded Price", "Highest Recorded Price", "Lowest Merchant Price", "Lowest In-Stock Price", "Average In-Stock Price", "Number of Merchants"]
        ws.append(headers)
        wb.save(filename)
        print(f"Created new workbook: {filename}")
        return wb

def calculate_price_metrics(offers):
    prices = [offer["price"] for offer in offers if "price" in offer]
    in_stock_prices = [offer["price"] for offer in offers if offer.get("availability") != "Out of Stock" and "price" in offer]
    
    lowest_price = min(prices) if prices else None
    lowest_in_stock_price = min(in_stock_prices) if in_stock_prices else None
    average_in_stock_price = sum(in_stock_prices) / len(in_stock_prices) if in_stock_prices else None
    number_of_offers = len(offers)
    
    return lowest_price, lowest_in_stock_price, average_in_stock_price, number_of_offers

def find_row_by_upc(ws, upc):
    upc_sheet_hyperlink = f'=HYPERLINK("#{upc}!A1", "{upc}")'
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[2] == upc_sheet_hyperlink:  # Check only the UPC for duplicates
            return idx
    return None

def update_or_add_to_master_sheet(wb, item, lowest_price, lowest_in_stock_price, average_in_stock_price,number_of_offers):
    ws = wb["Master"]
    upc = item.get("upc")
    row_to_update = find_row_by_upc(ws, upc)
    upc_sheet_hyperlink = f'=HYPERLINK("#{upc}!A1", "{upc}")'
    details = [item.get("ean"), item.get("title"), upc_sheet_hyperlink, item.get("gtin"), item.get("asin"), item.get("description"), item.get("brand"), item.get("model"), item.get("dimension"), item.get("weight"), item.get("category"), item.get("currency"), item.get("lowest_recorded_price"), item.get("highest_recorded_price"), lowest_price, lowest_in_stock_price, average_in_stock_price, number_of_offers]
    if row_to_update:
        tqdm.write(f"WARNING: Updating existing row for UPC: {upc}")
        for col, detail in enumerate(details, start=1):
            ws.cell(row=row_to_update, column=col, value=detail)
    else:
        tqdm.write(f"Adding new row for UPC: {upc}")
        ws.append(details)

# Function to create or append to UPC sheet, with updates to avoid duplicates
def add_to_upc_sheet(wb, upc, offers):
    if not upc or not offers:
        tqdm.write(f"Skipping empty UPC: {upc}")
        return
    if upc in wb.sheetnames:
        ws = wb[upc]
    else:
        ws = wb.create_sheet(upc)
        headers = ["Merchant", "Domain", "Title", "Currency", "List Price", "Price", "Shipping", "Condition", "Availability", "Link", "Updated"]
        ws.append(headers)
    existing_offers = {(row[0].value, row[2].value): row for row in ws.iter_rows(min_row=2)}
    for offer in offers:
        key = (offer.get("merchant"), offer.get("title"))
        details = [offer.get("merchant"), offer.get("domain"), offer.get("title"), offer.get("currency"), offer.get("list_price"), offer.get("price"), offer.get("shipping"), offer.get("condition"), offer.get("availability"), offer.get("link"), offer.get("updated_t")]
        if key in existing_offers:
            for col_num, detail in enumerate(details, start=1):
                existing_offers[key][col_num-1].value = detail
        else:
            ws.append(details)

# Modified process_upc function to include activity prints and handle rate limiting
def process_upc_enhanced(filename, upc):
    user_key = load_user_key()
    if not user_key: 
        url = f"https://api.upcitemdb.com/prod/trial/lookup?upc={upc}"
        headers = {}
    else: 
        url = f"https://api.upcitemdb.com/prod/v1/lookup?upc={upc}"
        headers = {"user_key": user_key}

    start_time = time.time()
    while True:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            if data.get("code") == "OK":
                wb = load_or_create_workbook(filename)
                try:
                    item = data["items"][0]  # Assuming one item per UPC for simplicity
                    try:
                        item.get("upc")
                        offers = item.get("offers", [])
                        lowest_price, lowest_in_stock_price, average_in_stock_price, number_of_offers = calculate_price_metrics(offers)
                        update_or_add_to_master_sheet(wb, item, lowest_price, lowest_in_stock_price, average_in_stock_price, number_of_offers)
                        add_to_upc_sheet(wb, upc, offers)
                        wb.save(filename)
                    except KeyError:
                        tqdm.write(f"Skipping invalid item for UPC: {upc}")
                        break
                except IndexError:
                    break
                break
            else:
                break
        else:
            # Retry if rate limit exceeded
            if response.status_code == 429:
                tqdm.write("API rate limit exceeded. Retrying in 10 seconds...")
                time.sleep(10)
                if time.time() - start_time > 58:
                    print("Exceeded maximum time limit. Exiting...")
                    exit()
            else:
                print("Error in response:", response.status_code)
                break

# Function to read UPCs from a TXT file or use a manual list
def read_upcs_or_use_default(filename, default_upcs):
    try:
        if os.path.exists(filename):
            print(f"Reading UPCs from {filename}")
            upcs = []
            with open(filename, mode='r') as file:
                for line in file:
                    upc = line.strip()
                    upcs.append(upc)
            return upcs if upcs else default_upcs
        else:
            print(f"{filename} file not found, using default UPC list")
            return default_upcs
    except Exception as e:
        print("Error reading file:", e)
        exit()

# After all operations, create a copy of the Master sheet into a new file
def save_master_sheet_separately(filename, master_only_filename):
    tqdm.write(f"Saving master sheet to {master_only_filename}")
    wb = load_workbook(filename)
    master_ws = wb["Master"]
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "master_upc_items"
    for row in master_ws.iter_rows(values_only=True):
        new_ws.append(row)
    new_wb.save(master_only_filename)

## Functions to call from outside this script
def lookup_request(upc):
    user_key = load_user_key()
    if not user_key: 
        url = f"https://api.upcitemdb.com/prod/trial/lookup?upc={upc}"
        headers = {}
    else: 
        url = f"https://api.upcitemdb.com/prod/v1/lookup?upc={upc}"
        headers = {"user_key": user_key}
    response = requests.get(url, headers=headers)  # Include headers in your request
    return response.json()

# Main function to process UPCs from a list
def main(outputfilename="upc_items.xlsx", masterfilename="master_upc_items.xlsx", input_filename="upc_lookup.txt", default_upcs=["887276550992"], skip_duplicates=True):
    upc_list = read_upcs_or_use_default(input_filename, default_upcs)
    print(f"Processing {len(upc_list)} UPCs")
    print(f"Output file: {outputfilename}")
    skip_duplicates = load_skip_duplicates()
    load_or_create_workbook(outputfilename)
    wb = load_workbook(outputfilename)
    if skip_duplicates == True:
        print("Skipping duplicates because flag is set to True")
    else:  
        print(f"Duplicates will not be skipped and will be overwritten in {outputfilename}")
    with tqdm(total=len(upc_list), position=0, leave=True) as pbar:
        for upc in upc_list:
            # if the upc is not a number, skip it
            if upc.isdigit():
                if skip_duplicates == True:
                    if find_row_by_upc(wb["Master"], upc):
                        tqdm.write(f"Skipping duplicate UPC: {upc}") 
                        pbar.update()
                        continue
                tqdm.write(f"Processing UPC: {upc}")
                process_upc_enhanced(outputfilename, upc)
                time.sleep(1)  # Respectful delay between API calls
                tqdm.write(f"Finished processing {upc}")
            else:
                tqdm.write(f"Skipping non-numeric UPC: {upc}")
            pbar.update()  # Update the progress bar after each UPC is processed or skipped
        save_master_sheet_separately(outputfilename, masterfilename)
        tqdm.write("Finished processing all UPCs")

if __name__ == "__main__":
    # Check if command-line arguments are provided
    if len(sys.argv) > 1:
        main(*sys.argv[1:])
    else:
        main()